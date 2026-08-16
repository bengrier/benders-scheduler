#!/usr/bin/env python3
"""Convert the league's Excel exports into the JSON the scheduler reads.

Usage:  python3 scripts/parse_excel.py

Reads  "C2 Benders Players.xlsx" and "C2 Schedule.xlsx" from the repo root,
writes data/players.json and data/schedule.json.

Re-run this whenever the league sends an updated schedule.
"""

import json
import re
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PLAYERS_XLSX = ROOT / "C2 Benders Players.xlsx"
SCHEDULE_XLSX = ROOT / "C2 Schedule.xlsx"
OUT_DIR = ROOT / "data"

OUR_TEAM = "Benders"

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}

# The season straddles a new year: Aug-Dec belong to the first year, Jan-Jul
# to the next. Pulled from the "2026 Winter C2 Division" title cell.
FALL_MONTHS = {8, 9, 10, 11, 12}


def clean(value):
    """Excel cells are littered with non-breaking spaces."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def parse_players():
    wb = openpyxl.load_workbook(PLAYERS_XLSX, data_only=True)
    ws = wb.worksheets[0]
    players = []
    for (cell,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        name = clean(cell)
        if not name:
            continue
        role = "skater"
        if re.search(r"\(G\)$", name):
            role = "goalie"
        elif re.search(r"\(C\)$", name):
            role = "captain"
        name = re.sub(r"\s*\((?:C|G|A)\)$", "", name).strip()
        first, _, last = name.partition(" ")
        players.append({
            "name": name,
            "first": first,
            "last": last or first,
            "role": role,
        })
    return players


def parse_season_year(title):
    match = re.search(r"(19|20)\d{2}", title)
    return int(match.group(0)) if match else date.today().year


def parse_date_cell(text, season_year):
    """'  Monday April 19  ' -> (date(2027, 4, 19), 'Monday')."""
    text = clean(text)
    if not text:
        return None, None
    weekday = None
    weekday_match = re.match(
        r"(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s+", text, re.I
    )
    if weekday_match:
        weekday = weekday_match.group(1).capitalize()
        text = text[weekday_match.end():]
    match = re.match(r"([A-Za-z]+)\s+(\d{1,2})", text)
    if not match:
        return None, None
    month = MONTHS.get(match.group(1).lower())
    if not month:
        return None, None
    year = season_year if month in FALL_MONTHS else season_year + 1
    day = date(year, month, int(match.group(2)))
    return day, weekday or day.strftime("%A")


def parse_matchup(text, default_time, default_rink):
    """'Blueberry Landesdogs vs Benders (9:30 PM)' -> structured game."""
    text = clean(text)
    if not text or text.lower() == "no game":
        return None

    time = default_time
    rink = default_rink
    override = re.search(r"\(([^)]*)\)\s*$", text)
    if override:
        inner = override.group(1)
        text = text[: override.start()].strip()
        time_match = re.search(r"\d{1,2}:\d{2}\s*[AP]M", inner, re.I)
        if time_match:
            time = time_match.group(0).upper().replace("  ", " ")
        rink_match = re.search(r"\b(West|East|North|South)\b", inner, re.I)
        if rink_match:
            rink = f"{rink_match.group(1).capitalize()} Rink"

    # Playoff rows look like "G4:Play-offs WG2 vs. WG3" -- keep them, but they
    # have no team names yet so we can't tell if we're playing.
    playoff = False
    label = None
    playoff_match = re.match(r"(G\d+)\s*:\s*Play-?offs?\s*(.*)", text, re.I)
    if playoff_match:
        playoff = True
        label = playoff_match.group(1).upper()
        text = playoff_match.group(2).strip()

    parts = re.split(r"\s+vs\.?\s+", text, maxsplit=1, flags=re.I)
    home = clean(parts[0])
    away = clean(parts[1]) if len(parts) > 1 else ""

    return {
        "home": home,
        "away": away,
        "time": time,
        "rink": rink,
        "playoff": playoff,
        "label": label,
        "isOurs": OUR_TEAM in (home, away),
    }


def parse_header_slot(text):
    """'7:00 PM - West Rink' -> ('7:00 PM', 'West Rink')."""
    text = clean(text)
    parts = [p.strip() for p in text.split("-", 1)]
    time = parts[0] if parts else ""
    rink = parts[1] if len(parts) > 1 else ""
    return time, rink


def parse_schedule():
    wb = openpyxl.load_workbook(SCHEDULE_XLSX, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    title = clean(rows[0][0])
    season_year = parse_season_year(title)
    slots = [parse_header_slot(c) for c in rows[1][1:]]

    games = []
    for row in rows[2:]:
        day, weekday = parse_date_cell(row[0], season_year)
        if not day:
            continue
        for index, cell in enumerate(row[1:]):
            if index >= len(slots):
                break
            default_time, default_rink = slots[index]
            game = parse_matchup(cell, default_time, default_rink)
            if not game:
                continue
            game["date"] = day.isoformat()
            game["weekday"] = weekday
            game["slot"] = index
            games.append(game)

    games.sort(key=lambda g: (g["date"], g["slot"]))
    return {"title": title, "team": OUR_TEAM, "games": games}


def main():
    OUT_DIR.mkdir(exist_ok=True)
    players = parse_players()
    schedule = parse_schedule()

    (OUT_DIR / "players.json").write_text(
        json.dumps(players, indent=2) + "\n", encoding="utf-8"
    )
    (OUT_DIR / "schedule.json").write_text(
        json.dumps(schedule, indent=2) + "\n", encoding="utf-8"
    )

    # Same payload as a plain script so index.html works when opened straight
    # from the filesystem, where fetch() of a local JSON file is blocked.
    (OUT_DIR / "data.js").write_text(
        "// Generated by scripts/parse_excel.py -- do not edit by hand.\n"
        f"window.PLAYERS = {json.dumps(players, indent=2)};\n"
        f"window.SCHEDULE = {json.dumps(schedule, indent=2)};\n",
        encoding="utf-8",
    )

    ours = sum(1 for g in schedule["games"] if g["isOurs"])
    playoffs = sum(1 for g in schedule["games"] if g["playoff"])
    print(f"players.json  {len(players)} players")
    print(f"schedule.json {len(schedule['games'])} games "
          f"({ours} Benders, {playoffs} playoff)")


if __name__ == "__main__":
    main()
